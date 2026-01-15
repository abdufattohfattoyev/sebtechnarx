import os
import re
import sqlite3
import asyncio
import traceback
from datetime import datetime

import pandas as pd
import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import InputFile
from aiogram.utils.exceptions import RetryAfter

from loader import dp, bot
from keyboards.default.knopkalar import admin_kb, cleanup_confirm_kb, back_kb
from data.config import ADMINS
from utils.api import api
from utils.db_api.database import (
    get_models, get_conn, DB_PATH, normalize_damage_format
)
from utils.db_api.user_database import (
    get_total_users, get_registered_users_count,
    get_active_users, get_global_stats, get_total_price_inquiries
)

# ================ KONSTANTALAR ================
MAX_FILE_SIZE_MB = 31
BATCH_SIZE = 100  # Har 100 qatorda commit
PROGRESS_UPDATE_INTERVAL = 500  # Har 500 qatorda progress yangilanadi


# ================ HOLATLAR ================

class ImportState(StatesGroup):
    waiting_file = State()


class CleanupState(StatesGroup):
    confirm = State()


# ================ YORDAMCHI FUNKSIYALAR ================

def get_total_prices_count():
    """Bazadagi jami narxlar soni"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM prices")
        count = c.fetchone()[0]
        conn.close()
        return count
    except:
        return 0


def get_prices_for_model(model_id):
    """Model uchun narxlarni olish"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM prices WHERE model_id=? LIMIT 10", (model_id,))
        rows = c.fetchall()
        conn.close()
        return rows
    except:
        return []


def clear_database():
    """Bazani tozalash (faqat narxlar va qismlar)"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM prices")
        c.execute("DELETE FROM parts")
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Bazani tozalashda xato: {e}")
        return False


async def safe_edit_message(progress_msg, text, parse_mode="HTML", max_retries=3):
    """Flood control bilan xabarni tahrirlash"""
    for retry in range(max_retries):
        try:
            await progress_msg.edit_text(text, parse_mode=parse_mode)
            return True
        except RetryAfter as e:
            await asyncio.sleep(e.timeout + 1)
        except Exception as e:
            if "message is not modified" in str(e).lower():
                return True
            if retry == max_retries - 1:
                print(f"Edit message xato: {e}")
                return False
            await asyncio.sleep(1)
    return False


async def safe_send_message(message, text, parse_mode="HTML", **kwargs):
    """Flood control bilan xabar yuborish"""
    max_retries = 3
    for retry in range(max_retries):
        try:
            return await message.answer(text, parse_mode=parse_mode, **kwargs)
        except RetryAfter as e:
            await asyncio.sleep(e.timeout + 1)
        except Exception as e:
            if retry == max_retries - 1:
                print(f"Send message xato: {e}")
                return None
            await asyncio.sleep(1)
    return None


def optimize_database_for_import():
    """Import uchun bazani optimallashtirish"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("PRAGMA journal_mode = WAL")
    c.execute("PRAGMA synchronous = NORMAL")
    c.execute("PRAGMA cache_size = 10000")
    c.execute("PRAGMA foreign_keys = OFF")
    c.execute("PRAGMA temp_store = MEMORY")
    conn.commit()
    return conn


def restore_database_settings(conn):
    """Bazani normal holatiga qaytarish"""
    try:
        c = conn.cursor()
        c.execute("PRAGMA foreign_keys = ON")
        c.execute("PRAGMA synchronous = FULL")
        conn.commit()
    except:
        pass


def get_cell_value(row, col_name, default=''):
    """DataFrame dan qiymat olish"""
    try:
        if col_name not in row.index:
            return str(default).strip()

        val = row[col_name]

        if hasattr(val, 'iloc'):
            val = val.iloc[0] if len(val) > 0 else default
        elif hasattr(val, '__iter__') and not isinstance(val, str):
            val_list = list(val)
            val = val_list[0] if len(val_list) > 0 else default

        result = str(val).strip()

        if result.lower() in ['nan', 'none', 'null', '']:
            return str(default).strip()

        return result
    except:
        return str(default).strip()


# ================ ADMIN HANDLERLAR ================

@dp.message_handler(lambda m: m.text == "🔧 Admin panel" and m.from_user.id in ADMINS)
async def admin_panel(message: types.Message):
    """Admin panel"""
    models_count = len(get_models())
    prices_count = get_total_prices_count()

    # Tariflar sonini olish
    tariffs_result = await api.get_tariffs()
    tariffs_count = len(tariffs_result.get('tariffs', [])) if tariffs_result.get('success') else 0

    text = f"""<b>👨‍💼 ADMIN PANEL</b>

📊 <b>Statistika:</b>
- 📱 Modellar: {models_count} ta
- 💰 Narxlar: {prices_count} ta
- 💳 Tariflar: {tariffs_count} ta
- 👥 Adminlar: {len(ADMINS)} ta

<b>🔧 Mavjud amallar:</b>
1. 📊 Statistika - To'liq statistika
2. 💳 Tariflar - Tariflarni ko'rish
3. 📥 Excel import - Faylni yuklash
4. 📤 Excel export - Bazani export qilish
5. 🧹 Tozalash - Test ma'lumotlarni o'chirish
6. 📱 Namuna - Model uchun narxlar

<b>⚠️ Diqqat:</b> Import dan oldin backup oling!
    """

    await message.answer(text, reply_markup=admin_kb(), parse_mode="HTML")


@dp.message_handler(lambda m: m.text == "📊 Statistika" and m.from_user.id in ADMINS)
async def admin_statistics(message: types.Message):
    """Admin statistika"""
    try:
        progress_msg = await message.answer("⏳ Statistika tayyorlanmoqda...")

        # Asosiy raqamlar
        models_count = len(get_models())
        prices_count = get_total_prices_count()
        total_users = get_total_users()
        registered_users = get_registered_users_count()

        # Faol foydalanuvchilar
        active_users_today = get_active_users(1)
        active_users_week = get_active_users(7)
        active_users_month = get_active_users(30)

        # Global statistika
        today_stats = get_global_stats('today')
        week_stats = get_global_stats('week')
        all_stats = get_global_stats('all')

        # Jami narxlatishlar
        total_price_inquiries = get_total_price_inquiries()

        text = f"""<b>📊 ADMIN STATISTIKA</b>

<b>👥 FOYDALANUVCHILAR:</b>
- Jami: <b>{total_users}</b> ta
- Ro'yxatdan o'tgan: <b>{registered_users}</b> ta
- Faol (bugun): <b>{active_users_today}</b> ta
- Faol (hafta): <b>{active_users_week}</b> ta
- Faol (oy): <b>{active_users_month}</b> ta

<b>📈 NARXLATISH:</b>
- Bugun: <b>{today_stats.get('total_inquiries', 0)}</b> ta
- Hafta: <b>{week_stats.get('total_inquiries', 0)}</b> ta
- Jami: <b>{total_price_inquiries}</b> ta

<b>📱 BAZA:</b>
- Modellar: <b>{models_count}</b> ta
- Narxlar: <b>{prices_count}</b> ta
"""

        # TOP modellar
        top_models = all_stats.get('top_models', [])[:3]
        if top_models:
            text += f"\n<b>🏆 TOP MODELLAR:</b>\n"
            for i, model in enumerate(top_models, 1):
                text += f"{i}. {model['model_name']}: <b>{model['count']}</b> marta\n"

        # Bugungi faol
        today_top_users = today_stats.get('top_users', [])[:3]
        if today_top_users:
            text += f"\n<b>⭐ BUGUNGI FAOL:</b>\n"
            for i, user in enumerate(today_top_users, 1):
                username = user.get('username') or user.get('first_name', 'User')
                display = f"@{username}" if user.get('username') else username
                text += f"{i}. {display}: <b>{user['count']}</b> ta\n"

        now = datetime.now()
        text += f"\n<b>🕐 Vaqt:</b> {now.strftime('%d.%m.%Y %H:%M')}"

        await progress_msg.edit_text(text, parse_mode="HTML")

    except Exception as e:
        await message.answer(f"❌ Xatolik:\n<code>{str(e)[:300]}</code>", parse_mode="HTML")
        traceback.print_exc()


@dp.message_handler(lambda m: m.text == "💳 Tariflar" and m.from_user.id in ADMINS)
async def show_tariffs(message: types.Message):
    """Tariflarni ko'rsatish"""
    try:
        result = await api.get_tariffs()

        if not result.get('success'):
            await message.answer("❌ Xatolik")
            return

        tariffs = result.get('tariffs', [])

        text = "<b>💳 TARIFLAR</b>\n\n"

        if tariffs:
            for tariff in tariffs:
                price_per_one = tariff['price'] / tariff['count']
                text += f"<b>{tariff['name']}</b>\n"
                text += f"  Soni: {tariff['count']} ta\n"
                text += f"  Narxi: {tariff['price']:,.0f} so'm\n"
                text += f"  Bitta: {price_per_one:,.0f} so'm\n\n"
        else:
            text += "❌ Tariflar yo'q\n\n"

        text += "<i>Tariflarni admin paneldan boshqaring</i>"

        await message.answer(text, parse_mode="HTML")

    except Exception as e:
        await message.answer(f"❌ Xatolik: {str(e)}")


@dp.message_handler(lambda m: m.text == "📤 Excel export" and m.from_user.id in ADMINS)
async def excel_export(message: types.Message):
    """Excel export - TO'G'RILANGAN SQL"""
    try:
        progress_msg = await message.answer("⏳ Tayyorlanmoqda...")

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # SIMPLE VA TO'G'RI QUERY
        query = """
        SELECT
            m.name as Model,
            p.storage_size as Xotira,
            CASE WHEN p.color_name IS NULL OR p.color_name = '' THEN 'Standart' ELSE p.color_name END as Rang,
            CASE WHEN p.sim_type = 'esim' THEN 'eSIM' ELSE 'SIM' END as SIM,
            p.battery_label as Batareya,
            CASE WHEN p.has_box = 1 THEN 'Bor' ELSE 'Yoq' END as Quti,
            p.damage_pct as Qismlar,
            p.price as Narx
        FROM prices p
        LEFT JOIN models m ON p.model_id = m.id
        WHERE m.id IS NOT NULL
        ORDER BY m.name, p.storage_size DESC
        """

        df = pd.read_sql_query(query, conn)
        conn.close()

        if df.empty:
            await progress_msg.edit_text("❌ Ma'lumot yo'q")
            return

        total = len(df)
        await progress_msg.edit_text(f"✅ {total:,} ta\n📊 Excel tayyorlanmoqda...")

        # Narxlarni formatlaymiz
        if 'Narx' in df.columns:
            df['Narx'] = df['Narx'].apply(lambda x: int(x) if pd.notna(x) else 0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"export_{message.from_user.id}_{timestamp}.xlsx"

        # Excel yozish
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Narxlar', index=False)

            # Excelni formatlashtirish
            workbook = writer.book
            worksheet = writer.sheets['Narxlar']

            # Header formatlashtirish
            from openpyxl.styles import Font, PatternFill, Alignment

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ustunlarni kenglashtirish va centerlashtirish
            for col_idx, column in enumerate(df.columns, 1):
                max_length = len(str(column))
                for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                adjusted_width = min(max_length + 2, 50)
                col_letter = worksheet.cell(1, col_idx).column_letter
                worksheet.column_dimensions[col_letter].width = adjusted_width

                # Barcha qatorlarni centerlashtirish
                for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        size_mb = os.path.getsize(file_path) / (1024 * 1024)

        # Faylni yuborish
        await message.answer_document(
            InputFile(file_path, filename=f"narxlar_{timestamp}.xlsx"),
            caption=f"📊 Jami: {total:,} ta\n📦 Hajm: {size_mb:.2f} MB\n✅ {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )

        # Cleanup
        os.remove(file_path)

        try:
            await progress_msg.delete()
        except:
            pass

    except Exception as e:
        error_text = str(e)[:500]
        print(f"❌ Export xatosi: {error_text}")
        await message.answer(f"❌ Xatolik:\n<code>{error_text}</code>", parse_mode="HTML")
        traceback.print_exc()


@dp.message_handler(lambda m: m.text == "❌ Bekor qilish", state=ImportState.waiting_file)
async def cancel_import(message: types.Message, state: FSMContext):
    """Bekor qilish"""
    await state.finish()
    await message.answer("✅ Bekor qilindi", reply_markup=admin_kb())


@dp.message_handler(content_types=['document'], state=ImportState.waiting_file)
async def process_import(message: types.Message, state: FSMContext):
    """Excel import - TO'LIQ TUZATILGAN VERSIYA"""
    user_id = message.from_user.id

    if user_id not in ADMINS:
        await state.finish()
        return

    # Fayl validatsiyasi
    file_name = message.document.file_name or "file"
    file_size_mb = message.document.file_size / (1024 * 1024)

    if not file_name.lower().endswith(('.xlsx', '.xls')):
        await message.answer(
            "❌ <b>Xato format!</b>\n\n"
            "Faqat .xlsx yoki .xls fayllarni qabul qilamiz.",
            parse_mode="HTML"
        )
        await state.finish()
        return

    if file_size_mb > 20:
        await message.answer(
            f"❌ <b>Fayl juda katta!</b>\n\n"
            f"Hajm: {file_size_mb:.1f} MB (max 20 MB)",
            parse_mode="HTML"
        )
        await state.finish()
        return

    progress_msg = await safe_send_message(message, text="⏳ Fayl yuklanmoqda...")
    if not progress_msg:
        await state.finish()
        return

    timestamp = int(datetime.now().timestamp())
    file_path = f"temp_{user_id}_{timestamp}.xlsx"
    conn = None

    try:
        # 1️⃣ FAYLNI YUKLASH
        try:
            file = await asyncio.wait_for(bot.get_file(message.document.file_id), timeout=30)
            await asyncio.wait_for(bot.download_file(file.file_path, file_path), timeout=90)
        except asyncio.TimeoutError:
            await safe_edit_message(progress_msg, "❌ Timeout - fayl yuklanmadi")
            await state.finish()
            return

        if not os.path.exists(file_path):
            await safe_edit_message(progress_msg, "❌ Fayl yuklanmadi")
            await state.finish()
            return

        actual_size = os.path.getsize(file_path) / (1024 * 1024)
        await safe_edit_message(progress_msg, f"✅ Yuklandi: {actual_size:.1f} MB\n📊 O'qilmoqda...")

        # 2️⃣ EXCEL O'QISH
        df = None
        try:
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
        except:
            try:
                df = pd.read_excel(file_path, dtype=str, engine='xlrd')
            except Exception as e:
                await safe_edit_message(progress_msg, f"❌ Excel o'qishda xato:\n{str(e)[:100]}")
                await state.finish()
                return

        if df is None or df.empty:
            await safe_edit_message(progress_msg, "❌ Excel faylida ma'lumot yo'q")
            await state.finish()
            return

        # 3️⃣ USTUNLARNI TAYYORLASH
        df = df.fillna('')
        df.columns = [str(col).strip() for col in df.columns]

        # 4️⃣ USTUNLARNI TOPISH VA MAPPING
        column_mapping = {
            'Model': ['model', 'iphone', 'samsung', 'xiaomi', 'redmi', 'phone'],
            'Xotira': ['xotira', 'storage', 'gb', 'память', 'size', 'ram'],
            'Rang': ['rang', 'color', 'цвет', 'reng'],
            'SIM': ['sim', 'sim turi', 'sim type', 'симкарта'],
            'Batareya': ['batareya', 'battery', 'batareyka', 'аккумулятор'],
            'Quti': ['quti', 'box', 'qora quti', 'коробка'],
            'Qismlar': ['qismlar', 'damage', 'holat', 'состояние', 'части', 'defects'],
            'Narx': ['narx', 'price', 'price_uzs', 'cena', 'стоимость', 'cost']
        }

        found_columns = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            for std_name, variants in column_mapping.items():
                if std_name not in found_columns:
                    if any(variant in col_lower for variant in variants):
                        found_columns[std_name] = col
                        break

        # Kerakli ustunlar tekshirish
        required_cols = ['Model', 'Xotira', 'Narx']
        missing = [col for col in required_cols if col not in found_columns]

        if missing:
            missing_text = "\n".join(f"• {col}" for col in missing)
            await safe_edit_message(
                progress_msg,
                f"❌ <b>Kerakli ustunlar topilmadi:</b>\n{missing_text}\n\n"
                f"<b>Topilgan ustunlar:</b>\n" +
                "\n".join(f"✅ {v} -> {k}" for k, v in found_columns.items())
            )
            await state.finish()
            return

        # Ustunlarni rename qilish
        rename_dict = {v: k for k, v in found_columns.items()}
        df = df.rename(columns=rename_dict)

        # Bo'sh qatorlarni o'chirish
        df = df[
            (df['Model'].astype(str).str.lower() != 'nan') &
            (df['Model'].astype(str).str.lower() != 'none') &
            (df['Model'].astype(str).str.strip() != '')
            ]

        total_rows = len(df)

        if total_rows == 0:
            await safe_edit_message(progress_msg, "❌ Excel faylida hech qanday to'liq ma'lumot yo'q")
            await state.finish()
            return

        await safe_edit_message(progress_msg, f"✅ {total_rows:,} ta qator topildi\n🔄 Yozilmoqda...\n0/{total_rows:,}")

        # 5️⃣ DATABASE OPTIMIZATSIYA
        conn = optimize_database_for_import()
        c = conn.cursor()
        conn.execute("BEGIN TRANSACTION")

        imported = 0
        skipped = 0
        batch_count = 0
        last_update = 0
        errors = []

        # 6️⃣ QA'TORLARNI QAYTA ISHLASH
        for index, row in df.iterrows():
            try:
                # Ma'lumotlarni olish
                model_name = get_cell_value(row, 'Model', '').strip()
                storage = get_cell_value(row, 'Xotira', '128GB').strip()
                color = get_cell_value(row, 'Rang', '').strip()
                sim_str = get_cell_value(row, 'SIM', '').strip()
                battery = get_cell_value(row, 'Batareya', '100%').strip()
                box_str = get_cell_value(row, 'Quti', '').strip()
                damage = get_cell_value(row, 'Qismlar', 'Yangi').strip()
                price_str = get_cell_value(row, 'Narx', '0').strip()

                # ✅ MODEL NOMINI TEKSHIRISH
                if not model_name or model_name.lower() in ['nan', 'none', 'null', 'undefined', '']:
                    skipped += 1
                    continue

                # ✅ NARXNI TOZALASH VA VALIDATSIYA
                price = 0.0
                try:
                    price_clean = re.sub(r'[^\d\.]', '', price_str)
                    if price_clean:
                        price = float(price_clean)
                    else:
                        skipped += 1
                        continue

                    if price <= 0:
                        skipped += 1
                        continue
                except (ValueError, TypeError):
                    skipped += 1
                    continue

                # ✅ SIM TURINI TEKSHIRISH
                sim_type = 'physical'
                if sim_str and 'esim' in sim_str.lower():
                    sim_type = 'esim'

                # ✅ QUTI HOLATINI TEKSHIRISH
                has_box = 1
                if box_str and any(w in box_str.lower() for w in ['yo\'q', 'yok', 'no', 'нет', 'none']):
                    has_box = 0

                # ✅ QISMLARNI NORMALIZATSIYA QILISH
                if damage and damage.lower() not in ['yangi', 'none', 'nan', '']:
                    damage = normalize_damage_format(damage)
                else:
                    damage = 'Yangi'

                # ✅ RANG TEKSHIRISH
                if color and color.lower() in ['nan', 'none', 'null', '']:
                    color = ''

                # ✅ MODEL QO'SHISH YOKI TOPISH
                c.execute("INSERT OR IGNORE INTO models (name) VALUES (?)", (model_name,))
                c.execute("SELECT id FROM models WHERE name = ?", (model_name,))
                model_result = c.fetchone()

                if not model_result:
                    skipped += 1
                    continue

                model_id = model_result[0]

                # ✅ XOTIRA QO'SHISH
                if storage and storage.lower() not in ['nan', 'none']:
                    c.execute("INSERT OR IGNORE INTO storages (model_id, size) VALUES (?, ?)",
                              (model_id, storage))

                # ✅ RANG QO'SHISH
                if color and color.lower() not in ['nan', 'none', '', 'standart']:
                    c.execute("INSERT OR IGNORE INTO colors (model_id, name) VALUES (?, ?)",
                              (model_id, color))

                # ✅ BATAREYA QO'SHISH
                if battery and battery.lower() not in ['nan', 'none']:
                    c.execute("INSERT OR IGNORE INTO batteries (model_id, label) VALUES (?, ?)",
                              (model_id, battery))

                # ✅ QISMLAR QO'SHISH
                if damage and damage != 'Yangi' and '+' in damage:
                    parts_list = [p.strip() for p in damage.split('+')]
                    for part in parts_list:
                        if part and part.lower() not in ['nan', 'none']:
                            c.execute("INSERT OR IGNORE INTO parts (model_id, part_name) VALUES (?, ?)",
                                      (model_id, part))

                # ✅ NARX QO'SHISH
                c.execute("""
                    INSERT OR REPLACE INTO prices
                    (model_id, storage_size, color_name, sim_type, battery_label, has_box, damage_pct, price, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (model_id, storage, color, sim_type, battery, has_box, damage, price, datetime.now()))

                imported += 1
                batch_count += 1

                # ✅ BATCH COMMIT (har 100 qatorda)
                if batch_count >= BATCH_SIZE:
                    conn.commit()
                    conn.execute("BEGIN TRANSACTION")
                    batch_count = 0

                # ✅ PROGRESS YANGILASH (har 500 qatorda)
                if imported % 500 == 0 or imported == total_rows:
                    progress_text = (
                        f"🔄 Yozilmoqda...\n\n"
                        f"✅ {imported:,}/{total_rows:,} ta\n"
                        f"⏭️ O'tkazildi: {skipped:,}"
                    )
                    await safe_edit_message(progress_msg, progress_text)
                    await asyncio.sleep(0.05)

            except Exception as row_error:
                skipped += 1
                error_msg = f"Qator {index + 2}: {str(row_error)[:50]}"
                errors.append(error_msg)
                print(f"❌ {error_msg}")

        # ✅ FINAL COMMIT
        try:
            conn.commit()
        except:
            pass

        # ✅ CLEANUP
        try:
            del df
        except:
            pass

        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass

        # ✅ NATIJA
        models_count = len(get_models())
        prices_count = get_total_prices_count()

        result_text = f"""✅ <b>IMPORT MUVAFFAQIYATLI!</b>

📊 <b>Natijalar:</b>
✅ Qo'shildi: <b>{imported:,}</b> ta narx
⏭️ O'tkazildi: <b>{skipped:,}</b> ta

📁 <b>Bazada jami:</b>
📱 Modellar: <b>{models_count}</b> ta
💰 Narxlar: <b>{prices_count:,}</b> ta

🕒 <b>Vaqt:</b> {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
"""

        if errors and len(errors) <= 5:
            result_text += f"\n<b>⚠️ Xatolar ({len(errors)}):</b>\n"
            for err in errors[:5]:
                result_text += f"• {err}\n"

        await safe_edit_message(progress_msg, result_text)

    except Exception as e:
        error_msg = f"❌ <b>Xatolik:</b>\n<code>{str(e)[:300]}</code>"
        await safe_edit_message(progress_msg, error_msg)
        print(f"❌ Import xatosi: {str(e)}")
        traceback.print_exc()

    finally:
        # ✅ CLEANUP
        if conn:
            try:
                restore_database_settings(conn)
                conn.close()
            except:
                pass

        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass

        await state.finish()
        await asyncio.sleep(2)
        await safe_send_message(message, text="🏠 Admin paneliga qaytish", reply_markup=admin_kb())


@dp.message_handler(lambda m: m.text == "📤 Excel export" and m.from_user.id in ADMINS)
async def excel_export(message: types.Message):
    """Excel export"""
    try:
        progress_msg = await message.answer("⏳ Tayyorlanmoqda...")

        conn = sqlite3.connect(DB_PATH)
        query = """
        SELECT
            m.name as Model,
            COALESCE(s.size, '') as Xotira,
            COALESCE(c.name, '') as Rang,
            CASE WHEN p.sim_type = 'esim' THEN 'eSIM' ELSE 'SIM' END as SIM,
            p.battery_label as Batareya,
            CASE WHEN p.has_box = 1 THEN 'Bor' ELSE 'Yo\'q' END as Quti,
            p.damage_pct as Qismlar,
            p.price as Narx
        FROM prices p
        LEFT JOIN models m ON p.model_id = m.id
        LEFT JOIN storages s ON p.model_id = s.model_id AND p.storage_size = s.size
        LEFT JOIN colors c ON p.model_id = c.model_id AND p.color_name = c.name
        ORDER BY m.name, s.size
        """

        df = pd.read_sql_query(query, conn)
        conn.close()

        if df.empty:
            await progress_msg.edit_text("❌ Ma'lumot yo'q")
            return

        total = len(df)
        await progress_msg.edit_text(f"✅ {total} ta\n📊 Excel...")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"export_{message.from_user.id}_{timestamp}.xlsx"

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Narxlar', index=False)

        size_mb = os.path.getsize(file_path) / (1024 * 1024)

        await message.answer_document(
            InputFile(file_path, filename=f"narxlar_{timestamp}.xlsx"),
            caption=f"📊 {total} ta • {size_mb:.1f} MB"
        )

        os.remove(file_path)
        await progress_msg.delete()

    except Exception as e:
        await message.answer(f"❌ Xatolik: {str(e)[:200]}")


@dp.message_handler(lambda m: m.text == "🧹 Bazani tozalash" and m.from_user.id in ADMINS)
async def cleanup_database(message: types.Message):
    """Tozalash"""
    models_count = len(get_models())
    prices_count = get_total_prices_count()

    text = f"""⚠️ <b>TOZALASH</b>

Joriy:
- Modellar: {models_count} ta
- Narxlar: {prices_count} ta

Barcha narxlar o'chadi!

Davom etasizmi?
    """

    await message.answer(text, reply_markup=cleanup_confirm_kb(), parse_mode="HTML")
    await CleanupState.confirm.set()


@dp.message_handler(state=CleanupState.confirm)
async def cleanup_confirm(message: types.Message, state: FSMContext):
    """Tasdiqlash"""
    if message.text == "✅ Ha, tozalash":
        try:
            success = clear_database()
            if success:
                await message.answer("✅ Tozalandi", reply_markup=admin_kb())
            else:
                await message.answer("❌ Xatolik", reply_markup=admin_kb())
        except Exception as e:
            await message.answer(f"❌ {str(e)}", reply_markup=admin_kb())
    else:
        await message.answer("✅ Bekor qilindi", reply_markup=admin_kb())

    await state.finish()


@dp.message_handler(lambda m: m.text == "📱 Namuna" and m.from_user.id in ADMINS)
async def sample_data(message: types.Message):
    """Namuna"""
    try:
        models = get_models()
        if not models:
            await message.answer("❌ Model yo'q")
            return

        model = models[0]
        prices = get_prices_for_model(model['id'])

        if not prices:
            await message.answer(f"❌ {model['name']} uchun narx yo'q")
            return

        text = f"<b>📱 {model['name']}</b>\n\n"
        for i, price in enumerate(prices[:5], 1):
            text += f"{i}. {price['storage_size']} - ${float(price['price']):,.0f}\n"

        text += f"\n📊 Jami: {len(prices)} ta"

        await message.answer(text, parse_mode="HTML")

    except Exception as e:
        await message.answer(f"❌ {str(e)}")


@dp.message_handler(lambda m: m.text == "📥 Excel import" and m.from_user.id in ADMINS)
async def import_start(message: types.Message):
    """Excel import boshlash"""
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton

    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.row(KeyboardButton("❌ Bekor qilish"), KeyboardButton("🏠 Bosh menyu"))

    text = """<b>📥 Excel faylni yuboring</b>

<b>✅ Talab qilinadi:</b>
• Format: .xlsx yoki .xls
• Hajm: maksimal 20 MB
• Kerakli ustunlar: Model, Xotira, Narx

<b>📊 Ixtiyoriy ustunlar:</b>
• Rang
• SIM (physical/esim)
• Batareya
• Quti (Bor/Yo'q)
• Qismlar (battery+screen+glass...)

<b>💡 Misol:</b>
Model | Xotira | Rang | SIM | Batareya | Quti | Qismlar | Narx
iPhone 15 | 128GB | Black | physical | 100% | Bor | Yangi | 5000000

<i>Faylni jo'natish uchun 📎 tugmasini bosing</i>"""

    await message.answer(text, reply_markup=kb, parse_mode="HTML")
    await ImportState.waiting_file.set()


@dp.message_handler(lambda m: m.text in ["❌ Bekor qilish", "🏠 Bosh menyu"], state=ImportState.waiting_file)
async def cancel_import(message: types.Message, state: FSMContext):
    """Import bekor qilish"""
    await state.finish()
    if message.text == "🏠 Bosh menyu":
        await message.answer("🏠 Bosh menyu", reply_markup=admin_kb())
    else:
        await message.answer("✅ Bekor qilindi", reply_markup=admin_kb())


@dp.message_handler(content_types=['document'], state=ImportState.waiting_file)
async def process_import(message: types.Message, state: FSMContext):
    """Excel import - YANGILANGAN"""
    user_id = message.from_user.id

    if user_id not in ADMINS:
        await state.finish()
        return

    # Fayl validatsiyasi
    file_name = message.document.file_name or "file"
    file_size_mb = message.document.file_size / (1024 * 1024)

    if not file_name.lower().endswith(('.xlsx', '.xls')):
        await message.answer(
            "❌ <b>Xato format!</b>\n\n"
            "Faqat .xlsx yoki .xls fayllarni qabul qilamiz.",
            parse_mode="HTML"
        )
        return

    if file_size_mb > 20:
        await message.answer(
            f"❌ <b>Fayl juda katta!</b>\n\n"
            f"Hajm: {file_size_mb:.1f} MB (max 20 MB)",
            parse_mode="HTML"
        )
        return

    progress_msg = await safe_send_message(message, text="⏳ Fayl yuklanmoqda...")
    if not progress_msg:
        await state.finish()
        return

    timestamp = int(datetime.now().timestamp())
    file_path = f"temp_{user_id}_{timestamp}.xlsx"
    conn = None

    try:
        # 1️⃣ FAYLNI YUKLASH
        try:
            file = await asyncio.wait_for(bot.get_file(message.document.file_id), timeout=30)
            await asyncio.wait_for(bot.download_file(file.file_path, file_path), timeout=90)
        except asyncio.TimeoutError:
            await safe_edit_message(progress_msg, "❌ Timeout - fayl yuklanmadi")
            await state.finish()
            return

        if not os.path.exists(file_path):
            await safe_edit_message(progress_msg, "❌ Fayl yuklanmadi")
            await state.finish()
            return

        await safe_edit_message(progress_msg, f"✅ Yuklandi: {file_size_mb:.1f} MB\n📊 O'qilmoqda...")

        # 2️⃣ EXCEL O'QISH
        try:
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
        except:
            try:
                df = pd.read_excel(file_path, dtype=str, engine='xlrd')
            except Exception as e:
                await safe_edit_message(progress_msg, f"❌ Excel o'qishda xato:\n{str(e)[:100]}")
                await state.finish()
                return

        df = df.fillna('')
        df.columns = [str(col).strip() for col in df.columns]

        # 3️⃣ USTUNLARNI TOPISH VA NORMALIZATSIYA
        column_mapping = {
            'Model': ['model', 'iphone', 'samsung', 'xiaomi', 'redmi'],
            'Xotira': ['xotira', 'storage', 'gb', 'память'],
            'Rang': ['rang', 'color', 'цвет', 'reng'],
            'SIM': ['sim', 'sim turi', 'sim type', 'симкарта'],
            'Batareya': ['batareya', 'battery', 'batareyka', 'аккумулятор'],
            'Quti': ['quti', 'box', 'qora quti', 'коробка'],
            'Qismlar': ['qismlar', 'damage', 'holat', 'состояние', 'части'],
            'Narx': ['narx', 'price', 'price_uzs', 'cena', 'стоимость']
        }

        found_columns = {}
        for col in df.columns:
            col_lower = col.lower()
            for std_name, variants in column_mapping.items():
                if any(v in col_lower for v in variants):
                    if std_name not in found_columns:
                        found_columns[std_name] = col
                    break

        # Kerakli ustunlar tekshirish
        required_cols = ['Model', 'Xotira', 'Narx']
        missing = [col for col in required_cols if col not in found_columns]

        if missing:
            await safe_edit_message(
                progress_msg,
                f"❌ <b>Ustunlar topilmadi:</b>\n" + "\n".join(f"• {col}" for col in missing)
            )
            await state.finish()
            return

        # Ustunlarni rename qilish
        rename_dict = {v: k for k, v in found_columns.items()}
        df = df.rename(columns=rename_dict)

        total_rows = len(df)
        await safe_edit_message(progress_msg, f"✅ {total_rows:,} ta qator topildi\n🔄 Yozilmoqda... 0/{total_rows:,}")

        # 4️⃣ DATABASE OPTIMIZATSIYA
        conn = optimize_database_for_import()
        c = conn.cursor()
        conn.execute("BEGIN TRANSACTION")

        imported = 0
        skipped = 0
        batch_count = 0
        last_update = 0

        # 5️⃣ QA'TORLARNI QAYTA ISHLASH
        for index, row in df.iterrows():
            try:
                model_name = get_cell_value(row, 'Model', '').strip()
                storage = get_cell_value(row, 'Xotira', '128GB').strip()
                color = get_cell_value(row, 'Rang', 'Standart').strip()
                sim_str = get_cell_value(row, 'SIM', 'physical').strip()
                battery = get_cell_value(row, 'Batareya', '100%').strip()
                box_str = get_cell_value(row, 'Quti', 'Bor').strip()
                damage = get_cell_value(row, 'Qismlar', 'Yangi').strip()
                price_str = get_cell_value(row, 'Narx', '0').strip()

                # Model nomini tekshirish
                if not model_name or model_name.lower() in ['nan', 'none', 'null', '']:
                    skipped += 1
                    continue

                # NARXNI TOZALASH
                price = 0.0
                try:
                    price_clean = re.sub(r'[^\d\.]', '', price_str)
                    if price_clean:
                        price = float(price_clean)
                    if price == 0:
                        skipped += 1
                        continue
                except:
                    skipped += 1
                    continue

                # SIM TURI
                sim_type = 'physical'
                if sim_str and 'esim' in sim_str.lower():
                    sim_type = 'esim'

                # QUTI
                has_box = 'Bor'
                if box_str and any(w in box_str.lower() for w in ['yo\'q', 'yok', 'no', 'нет']):
                    has_box = 'Yo\'q'

                # QISMLAR NORMALIZATSIYA
                damage = normalize_damage_format(damage)

                # MODEL QO'SHISH
                c.execute("INSERT OR IGNORE INTO models (name) VALUES (?)", (model_name,))
                c.execute("SELECT id FROM models WHERE name = ?", (model_name,))
                model_result = c.fetchone()

                if not model_result:
                    skipped += 1
                    continue

                model_id = model_result[0]

                # XOTIRA QO'SHISH
                if storage:
                    c.execute("INSERT OR IGNORE INTO storages (model_id, size) VALUES (?, ?)",
                              (model_id, storage))

                # RANG QO'SHISH
                if color and color.lower() not in ['nan', 'none', '', 'standart']:
                    c.execute("INSERT OR IGNORE INTO colors (model_id, name) VALUES (?, ?)",
                              (model_id, color))

                # BATAREYA QO'SHISH
                if battery:
                    c.execute("INSERT OR IGNORE INTO batteries (model_id, label) VALUES (?, ?)",
                              (model_id, battery))

                # QISMLAR QO'SHISH
                if damage and damage != "Yangi" and '+' in damage:
                    parts = [p.strip() for p in damage.split('+')]
                    for part in parts:
                        if part:
                            c.execute("INSERT OR IGNORE INTO parts (model_id, part_name) VALUES (?, ?)",
                                      (model_id, part))

                # NARX QO'SHISH
                color_for_price = color if color and color.lower() != 'standart' else ""
                has_box_int = 1 if has_box == 'Bor' else 0

                c.execute("""
                    INSERT OR REPLACE INTO prices
                    (model_id, storage_size, color_name, sim_type, battery_label, has_box, damage_pct, price)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (model_id, storage, color_for_price, sim_type, battery, has_box_int, damage, price))

                imported += 1
                batch_count += 1

                # BATCH COMMIT
                if batch_count >= BATCH_SIZE:
                    conn.commit()
                    conn.execute("BEGIN TRANSACTION")
                    batch_count = 0

                # PROGRESS YANGILASH
                if imported - last_update >= PROGRESS_UPDATE_INTERVAL or imported == total_rows:
                    progress_text = f"🔄 Yozilmoqda...\n\n✅ {imported:,}/{total_rows:,} ta\n⏭️ O'tkazildi: {skipped:,}"
                    await safe_edit_message(progress_msg, progress_text)
                    last_update = imported
                    await asyncio.sleep(0.1)

            except Exception as row_error:
                skipped += 1
                print(f"Qator {index + 2} xato: {row_error}")

        # FINAL COMMIT
        conn.commit()

        # CLEANUP
        try:
            del df
            os.remove(file_path)
        except:
            pass

        # NATIJA
        models_count = len(get_models())
        prices_count = get_total_prices_count()

        result_text = f"""✅ <b>IMPORT YAKUNLANDI!</b>

<b>📊 Natijalar:</b>
✅ Qo'shildi: <b>{imported:,}</b> ta
⏭️ O'tkazildi: <b>{skipped:,}</b> ta

<b>📁 Bazada:</b>
📱 Modellar: <b>{models_count}</b> ta
💰 Narxlar: <b>{prices_count:,}</b> ta

🕒 {datetime.now().strftime('%d.%m.%Y %H:%M')}
"""

        await safe_edit_message(progress_msg, result_text)

    except Exception as e:
        error_msg = f"❌ <b>Xatolik:</b>\n<code>{str(e)[:300]}</code>"
        await safe_edit_message(progress_msg, error_msg)
        traceback.print_exc()

    finally:
        if conn:
            try:
                restore_database_settings(conn)
                conn.close()
            except:
                pass

        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass

        await state.finish()
        await asyncio.sleep(2)
        await safe_send_message(message, text="🏠 Admin paneliga qaytish", reply_markup=admin_kb())